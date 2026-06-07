"""
GMP Automation System - Excel Generator
Generates formatted Excel files for all 5 test types.
"""

import os
import math
from copy import copy as _copy_obj
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import MultiLevelStrRef as _MultiLevelStrRef
from openpyxl.utils import get_column_letter

# openpyxl serialises CT_MultiLvlStrRef with <multiLvlStrCache> before <f>, but
# the OOXML schema requires <f> first. The wrong order makes Excel flag the file
# as needing repair (and on macOS it can drop the multi-level X-axis entirely).
# Reorder the element sequence once, globally, so the generated XML is valid.
from config import (
    AIRBORNE_PARTICLE, AIR_VELOCITY, AIR_CHANGE_RATE, HEPA_FILTER, AIRFLOW_PATTERN,
    ALERT_FILL_RED, HEADER_FILL, WHITE_FILL,
    HEADER_FONT, TITLE_FONT, DATA_FONT,
    CENTER_ALIGN, LEFT_ALIGN, THIN_BORDER,
    get_semester_label, semester_sort_key,
    OUTPUT_FOLDER
)

# ── Limit line color map (RRGGBB for openpyxl) ────────────────────────────────
_LIMIT_COLORS = {
    'A': {'alert': 'C00000', 'action': 'FF3300'},
    'B': {'alert': 'C07000', 'action': 'FF9900'},
    'C': {'alert': '375623', 'action': '70AD47'},
    'D': {'alert': '7030A0', 'action': '9B59B6'},
    'lower': 'C00000',
    'upper': 'C07000',
    'limit': 'C00000',
}


def _safe_float(value):
    if value is None or value == '':
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _unique_ordered(iterable):
    seen = set(); result = []
    for x in iterable:
        if x not in seen:
            seen.add(x); result.append(x)
    return result


def _nice_y_max(raw_max):
    """Round raw_max up to a clean chart Y-axis ceiling."""
    if raw_max <= 0:
        return 10
    magnitude = 10 ** math.floor(math.log10(raw_max))
    return math.ceil(raw_max / magnitude) * magnitude


def _style_limit_series(series, color_hex, is_action=False):
    """Apply dashed colored line style to a chart series."""
    try:
        line = series.graphicalProperties.line
        line.solidFill = color_hex
        line.w = 19050          # 1.5 pt in EMU
        line.prstDash = 'sysDash' if is_action else 'dash'
    except Exception:
        pass
    try:
        series.marker.symbol = 'none'
    except Exception:
        pass
    try:
        series.smooth = False
    except Exception:
        pass


def _write_limit_columns(ws, limit_specs, header_row, data_start_row, data_end_row, start_col):
    """
    Write constant limit values into worksheet columns.
    limit_specs: list of (label, value)
    Returns the column index AFTER the last limit column.
    """
    for i, (label, value) in enumerate(limit_specs):
        col = start_col + i
        ws.cell(row=header_row, column=col, value=label)
        for r in range(data_start_row, data_end_row + 1):
            ws.cell(row=r, column=col, value=value)
    return start_col + len(limit_specs)


def _build_linechart_for_limits(ws, limit_specs, colors, header_row,
                                 data_start_row, data_end_row, start_col):
    """
    Create a LineChart with one dashed series per limit spec.
    limit_specs: list of (label, value)
    colors:      list of (color_hex, is_action) matching limit_specs
    """
    if not limit_specs:
        return None
    line = LineChart()
    line.grouping = 'standard'
    for i, (label, value) in enumerate(limit_specs):
        col = start_col + i
        ref = Reference(ws, min_col=col, min_row=header_row, max_row=data_end_row)
        line.add_data(ref, titles_from_data=True)
        s = line.series[-1]
        _style_limit_series(s, colors[i][0], colors[i][1])
        # Limit lines are identified in the right-side legend, not via in-plot
        # data labels (those collide with the bars and with each other).
    return line


def _write_label_column(ws, data_end_row, room_num_col, name_col, label_col):
    """Write a combined 'room_num  name' label column used as the chart X-axis."""
    ws.cell(row=1, column=label_col, value='실명')
    for r in range(2, data_end_row + 1):
        num  = ws.cell(row=r, column=room_num_col).value
        name = ws.cell(row=r, column=name_col).value
        label = f"{num}  {name}" if num is not None else str(name or '')
        ws.cell(row=r, column=label_col, value=label)


def _set_str_categories(chart, ws, label_col, data_end_row):
    from openpyxl.chart.data_source import (
        AxDataSource,
        StrRef,
        StrData,
        StrVal,
    )

    col_letter = get_column_letter(label_col)

    formula = (
        f"'{ws.title}'!"
        f"${col_letter}$2:${col_letter}${data_end_row}"
    )

    pts = []

    for idx, r in enumerate(range(2, data_end_row + 1)):
        pts.append(
            StrVal(
                idx=idx,
                v=str(
                    ws.cell(
                        row=r,
                        column=label_col
                    ).value or ""
                )
            )
        )

    cache = StrData(
        ptCount=len(pts),
        pt=pts
    )

    cat = AxDataSource(
        strRef=StrRef(
            f=formula,
            strCache=cache
        )
    )

    for s in chart.series:
        s.cat = cat

def _add_last_point_label(series, n_points):
    """Show series name as a data label on the LAST data point only (positioned right)."""
    try:
        from openpyxl.chart.label import DataLabel, DataLabelList
        last_idx = max(n_points - 1, 0)
        last_label = DataLabel(
            idx=last_idx,
            dLblPos='r',
            showSerName=True,
            showVal=False,
            showCatName=False,
            showLegendKey=False,
            showPercent=False,
            showBubbleSize=False,
        )
        series.dLbls = DataLabelList(
            dLbl=[last_label],
            showLegendKey=False,
            showVal=False,
            showCatName=False,
            showSerName=False,
            showPercent=False,
            showBubbleSize=False,
        )
    except Exception:
        pass


def _hide_limit_lines_from_legend(chart, n_bar_series, n_limit_series):
    """Remove limit line series from the chart legend."""
    try:
        from openpyxl.chart.legend import LegendEntry
        if chart.legend is None:
            return
        for i in range(n_bar_series, n_bar_series + n_limit_series):
            chart.legend.legendEntry.append(LegendEntry(idx=i, delete=True))
    except Exception:
        pass


_NOTE_TEXT = (
    "참고: 위 차트에서 특정 등급의 제한선이 표시되지 않는 경우, "
    "해당 등급의 모든 측정 결과가 아직 기준을 만족하고 있음을 의미합니다. "
    "해당 등급의 제한선은 제한값이 매우 높아 차트 표시 범위를 초과하므로 "
    "나타나지 않습니다."
)


def _make_chart_sheet(wb, sheet_name, chart_title,
                       cat_col_specs, data_rows_map, semesters,
                       limit_specs, limit_colors,
                       y_num_fmt=None, note_text=_NOTE_TEXT):
    """
    Generic chart sheet builder.

    cat_col_specs : list of (header, key) for category columns.
                    Last column is used as chart X-axis label.
    data_rows_map : list of dicts with keys matching cat_col_specs keys + 'semester' + 'value'
    semesters     : list of semester labels (sorted)
    limit_specs   : list of (label, value) — auto-filtered to visible range
    limit_colors  : list of (color_hex, is_action) matching limit_specs
    """
    ws = wb.create_sheet(title=sheet_name)
    n_cat = len(cat_col_specs)
    n_sems = len(semesters)

    # ── Unique ordered categories ─────────────────────────────────────────────
    cat_keys = [k for _, k in cat_col_specs]
    unique_cats = _unique_ordered(
        tuple(row[k] for k in cat_keys) for row in data_rows_map
    )

    # ── Write header row ──────────────────────────────────────────────────────
    for ci, (hdr, _) in enumerate(cat_col_specs):
        ws.cell(row=1, column=1 + ci, value=hdr)
    for si, sem in enumerate(semesters):
        ws.cell(row=1, column=1 + n_cat + si, value=sem)

    # ── Write data rows ───────────────────────────────────────────────────────
    for ri, cat_vals in enumerate(unique_cats):
        row = 2 + ri
        for ci, val in enumerate(cat_vals):
            ws.cell(row=row, column=1 + ci, value=val)
        for si, sem in enumerate(semesters):
            matched = next(
                (r['value'] for r in data_rows_map
                 if all(r[k] == cat_vals[ki] for ki, (_, k) in enumerate(cat_col_specs))
                 and r['semester'] == sem),
                None
            )
            cell = ws.cell(row=row, column=1 + n_cat + si, value=matched)
            if y_num_fmt and matched is not None:
                cell.number_format = y_num_fmt

    n_items = len(unique_cats)
    data_end_row = 1 + n_items

    # ── Auto-scale: compute y_max from data, filter visible limits ────────────
    data_vals = [
        ws.cell(row=r, column=1 + n_cat + si).value
        for r in range(2, data_end_row + 1)
        for si in range(n_sems)
        if ws.cell(row=r, column=1 + n_cat + si).value is not None
    ]
    data_max = max(data_vals or [1])
    # Y-axis is driven by the DATA so the bars stay tall and readable.
    # A limit line is only drawn if it falls within this visible range; limits
    # far above the data (e.g. Grade D) would otherwise blow up the Y-axis and
    # squash every bar to the bottom of the plot.
    y_max = _nice_y_max(data_max * 1.05)
    vis_specs  = [(l, v) for (l, v), c in zip(limit_specs, limit_colors) if v <= y_max]
    vis_colors = [c       for (l, v), c in zip(limit_specs, limit_colors) if v <= y_max]

    # ── X-axis labels from name column (Col C for 3-level category tables) ──
    # Requested behavior: use the values from column C starting at row 2.
    # For sheets with fewer category columns, use the last category column.
    label_col = n_cat

    limit_start_col = 1 + n_cat + n_sems

    # ── Write VISIBLE limit columns only ──────────────────────────────────────
    _write_limit_columns(ws, vis_specs, 1, 2, data_end_row, limit_start_col)

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci in range(n_cat):
        widths = [8, 8, 22, 22]
        ws.column_dimensions[get_column_letter(1 + ci)].width = widths[ci] if ci < len(widths) else 12
    for si in range(n_sems):
        ws.column_dimensions[get_column_letter(1 + n_cat + si)].width = 14
    for i in range(len(vis_specs)):
        ws.column_dimensions[get_column_letter(limit_start_col + i)].width = 14

    # Keep the first visible limit/data range readable for users opening the sheet.
    for col in range(5, min(10, limit_start_col + max(len(vis_specs), 1) - 1) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width,
            14,
        )

    # ── Bar chart ─────────────────────────────────────────────────────────────
    bar = BarChart()
    try:
        bar.y_axis.majorGridlines = None
    except:
        pass
    bar.height = 20
    bar.title = chart_title
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.type = 'col'; bar.grouping = 'clustered'
    bar.title = chart_title
    bar.x_axis.title = "측정 위치"
    # bar.y_axis.title = "측정값"
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = y_max
    bar.width = 36; bar.height = 18
    # Legend on the right so it sits in whitespace and never overlaps the plot.
    bar.legend.position = 'r'
    bar.legend.overlay = False

    try:
        bar.x_axis.txPr = None
    except:
        pass

    bar.x_axis.tickLblSkip = 1

    if y_num_fmt:
        bar.y_axis.numFmt = y_num_fmt

    for si in range(n_sems):
        col = 1 + n_cat + si
        ref = Reference(ws, min_col=col, min_row=1, max_row=data_end_row)
        bar.add_data(ref, titles_from_data=True)

    cats = Reference(
        ws,
        min_col=label_col,
        min_row=2,
        max_row=data_end_row
    )

    bar.set_categories(cats)

    # ── Limit line chart ──────────────────────────────────────────────────────
    if vis_specs:
        line = _build_linechart_for_limits(
            ws, vis_specs, vis_colors, 1, 2, data_end_row, limit_start_col
        )
        if line:
            bar += line

    bar.x_axis.axPos = "b"
    bar.x_axis.tickLblPos = "nextTo"

    # ── Anchor chart to the right of the table ────────────────────────────────
    chart_col = get_column_letter(limit_start_col + len(vis_specs) + 1)
    ws.add_chart(bar, f"{chart_col}1")

    # ── Note text ─────────────────────────────────────────────────────────────
    if note_text:
        note_row = data_end_row + 2
        note_cell = ws.cell(row=note_row, column=1, value=note_text)
        from openpyxl.styles import Alignment
        note_cell.alignment = Alignment(wrap_text=True, vertical='top')
        merge_end = max(n_cat + n_sems, 3)
        try:
            ws.merge_cells(start_row=note_row, start_column=1,
                            end_row=note_row + 1, end_column=merge_end)
        except Exception:
            pass
        ws.row_dimensions[note_row].height = 45

    return ws


def apply_cell_style(cell, font=DATA_FONT, alignment=CENTER_ALIGN, border=THIN_BORDER, fill=None):
    """Apply styling to a cell."""
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill


def apply_header_style(ws, row, col_start, col_end):
    """Apply header styling to a row range."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)


def create_clustered_chart(ws, data_ws, title, categories_col, values_cols, semester_labels,
                           chart_position, y_max, y_title="", data_start_row=2, data_end_row=None,
                           chart_width=30, chart_height=15):
    """Create a clustered column chart."""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = y_title
    chart.y_axis.scaling.max = y_max
    chart.y_axis.scaling.min = 0
    chart.width = chart_width
    chart.height = chart_height
    chart.legend.position = 'r'

    if data_end_row is None:
        data_end_row = data_ws.max_row

    # Categories (x-axis labels)
    cats = Reference(data_ws, min_col=categories_col, min_row=data_start_row,
                     max_row=data_end_row)
    chart.set_categories(cats)

    # Data series
    for col_idx, label in zip(values_cols, semester_labels):
        values = Reference(data_ws, min_col=col_idx, min_row=data_start_row - 1,
                          max_row=data_end_row)
        chart.add_data(values, titles_from_data=True)

    ws.add_chart(chart, chart_position)
    return chart


# =============================================================================
# A. AIRBORNE PARTICLE TEST EXCEL GENERATOR
# =============================================================================

def generate_airborne_particle_excel(all_ahu_data, output_path=None):
    """
    Generate Airborne Particle Test Result and Graph Excel file.

    all_ahu_data: dict keyed by AHU number, each value is a list of semester data:
    {
        "33": [
            {
                "semester": "2025 (하)",
                "date": "2025.08.14",
                "rooms": [
                    {
                        "grade": "B", "room_number": "2142", "room_name": "무균 실험실",
                        "measurements": [{"point": 1, "value_05": 121, "value_50": 7}, ...]
                    }, ...
                ]
            }, ...
        ]
    }
    """
    if output_path is None:
        output_path = os.path.join(OUTPUT_FOLDER, AIRBORNE_PARTICLE['excel_filename'])
    wb = Workbook()
    wb.remove(wb.active)
    for ahu_num in sorted(all_ahu_data.keys(), key=lambda x: int(x) if x.isdigit() else x):
        ahu_semesters = all_ahu_data[ahu_num]
        ahu_semesters.sort(key=lambda s: semester_sort_key(s['semester']))
        _create_airborne_data_sheet(wb, ahu_num, ahu_semesters)
        table_ws = _create_airborne_table_sheet(wb, ahu_num, ahu_semesters)
        _create_airborne_chart_sheet(wb, ahu_num, table_ws, '0.5')
        _create_airborne_chart_sheet(wb, ahu_num, table_ws, '5.0')
    wb.save(output_path)
    return output_path


def _create_airborne_data_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Data sheet for Airborne Particle Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Data")

    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = '부유입자 측정 기록서'
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN

    # Headers (row 7-8)
    headers_row7 = ['NO', '청정 등급', '실번호', '실명', '측정번호', '측정값', '', '', '', '측정일자']
    headers_row8 = ['', '', '', '', '', '0.5 µm', 'Average 0.5㎛', '5.0 µm', 'Average 5.0㎛', '']

    for col_idx, val in enumerate(headers_row7, 1):
        cell = ws.cell(row=7, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    for col_idx, val in enumerate(headers_row8, 1):
        cell = ws.cell(row=8, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    # Merge header cells
    ws.merge_cells('F7:I7')  # 측정값 spans 4 columns
    ws.merge_cells('A7:A8')
    ws.merge_cells('B7:B8')
    ws.merge_cells('C7:C8')
    ws.merge_cells('D7:D8')
    ws.merge_cells('E7:E8')
    ws.merge_cells('J7:J8')

    # Data rows
    current_row = 9
    row_num = 1

    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        rooms = sem_data['rooms']

        for room in rooms:
            grade = room['grade']
            room_number = room['room_number']
            room_name = room['room_name']
            measurements = room['measurements']
            n_points = len(measurements)

            # Calculate averages
            avg_05 = sum(m['value_05'] for m in measurements) / n_points if n_points > 0 else 0
            avg_50 = sum(m['value_50'] for m in measurements) / n_points if n_points > 0 else 0
            avg_05_rounded = round(avg_05)
            avg_50_rounded = round(avg_50, 2)

            start_row = current_row

            for i, m in enumerate(measurements):
                r = current_row + i
                ws.cell(row=r, column=1, value=row_num + i)  # NO
                apply_cell_style(ws.cell(row=r, column=1))

                if i == 0:
                    ws.cell(row=r, column=2, value=grade)
                    ws.cell(row=r, column=3, value=int(room_number) if room_number.isdigit() else room_number)
                    ws.cell(row=r, column=4, value=room_name)

                for c in range(2, 5):
                    apply_cell_style(ws.cell(row=r, column=c))

                ws.cell(row=r, column=5, value=m['point'])
                apply_cell_style(ws.cell(row=r, column=5))

                # 0.5 µm value
                cell_05 = ws.cell(row=r, column=6, value=m['value_05'])
                apply_cell_style(cell_05)
                # Conditional formatting for 0.5 µm
                if grade in AIRBORNE_PARTICLE['alert_limits']['0.5']:
                    limit = AIRBORNE_PARTICLE['alert_limits']['0.5'][grade]
                    if m['value_05'] > limit:
                        cell_05.fill = ALERT_FILL_RED

                # 5.0 µm value
                cell_50 = ws.cell(row=r, column=8, value=m['value_50'])
                apply_cell_style(cell_50)
                # Conditional formatting for 5.0 µm
                if grade in AIRBORNE_PARTICLE['alert_limits']['5.0']:
                    limit = AIRBORNE_PARTICLE['alert_limits']['5.0'][grade]
                    if m['value_50'] > limit:
                        cell_50.fill = ALERT_FILL_RED

                # Apply border to Average and 측정일자 columns
                apply_cell_style(ws.cell(row=r, column=7))
                apply_cell_style(ws.cell(row=r, column=9))
                apply_cell_style(ws.cell(row=r, column=10))

            end_row = current_row + n_points - 1

            # Merge cells for grade, room_number, room_name
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)

            # Average 0.5 µm (merged cell with AVERAGE formula)
            avg_cell_05 = ws.cell(row=start_row, column=7)
            avg_cell_05.value = f"=AVERAGE(F{start_row}:F{end_row})"
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

            # Average 5.0 µm (merged cell with AVERAGE formula)
            avg_cell_50 = ws.cell(row=start_row, column=9)
            avg_cell_50.value = f"=AVERAGE(H{start_row}:H{end_row})"
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9)

            # 측정일자 (merged cell)
            date_cell = ws.cell(row=start_row, column=10, value=semester_label)
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)

            row_num += n_points
            current_row = end_row + 1

    # Set column widths
    widths = {'A': 6, 'B': 10, 'C': 10, 'D': 22, 'E': 10, 'F': 12, 'G': 14, 'H': 12, 'I': 14, 'J': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_airborne_table_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Table sheet for Airborne Particle Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Table")

    # Headers (row 1)
    headers = ['NO', '청정 등급', '실번호', '실명', 'Average 0.5㎛', 'Average 5.0㎛', '측정일자']
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    row = 2
    no = 1
    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        for room in sem_data['rooms']:
            measurements = room['measurements']
            n_points = len(measurements)
            avg_05 = sum(m['value_05'] for m in measurements) / n_points if n_points > 0 else 0
            avg_50 = sum(m['value_50'] for m in measurements) / n_points if n_points > 0 else 0

            ws.cell(row=row, column=1, value=no)
            ws.cell(row=row, column=2, value=room['grade'])
            ws.cell(row=row, column=3, value=int(room['room_number']) if room['room_number'].isdigit() else room['room_number'])
            ws.cell(row=row, column=4, value=room['room_name'])
            ws.cell(row=row, column=5, value=round(avg_05))
            ws.cell(row=row, column=6, value=avg_50)
            ws.cell(row=row, column=7, value=semester_label)

            for c in range(1, 8):
                apply_cell_style(ws.cell(row=row, column=c))

            no += 1
            row += 1

    widths = {'A': 6, 'B': 10, 'C': 10, 'D': 22, 'E': 16, 'F': 16, 'G': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_airborne_chart_sheet(wb, ahu_num, table_ws, particle_size):
    """Create AHU-X chart sheet for Airborne Particle Test with multi-level X-axis and limit lines."""
    ps_key = '0.5' if particle_size == '0.5' else '5.0'
    value_col = 5 if particle_size == '0.5' else 6

    # Read data from Table sheet
    data_rows = []
    for r in range(2, table_ws.max_row + 1):
        grade    = str(table_ws.cell(row=r, column=2).value or '')
        room_num = table_ws.cell(row=r, column=3).value
        name     = table_ws.cell(row=r, column=4).value
        value    = table_ws.cell(row=r, column=value_col).value
        semester = table_ws.cell(row=r, column=7).value
        if name is None:
            continue
        data_rows.append({'grade': grade, 'room_num': room_num, 'name': name,
                           'value': value, 'semester': semester})

    semesters = sorted({d['semester'] for d in data_rows if d['semester']}, key=semester_sort_key)
    grades_present = sorted({d['grade'] for d in data_rows if d['grade']})

    alert_map  = AIRBORNE_PARTICLE['alert_limits'][ps_key]
    action_map = AIRBORNE_PARTICLE.get('action_limits', {}).get(ps_key, {})
    limit_specs, colors = [], []
    for g in grades_present:
        if g in alert_map:
            limit_specs.append((f"Grade {g} 경고기준: {alert_map[g]:,}", alert_map[g]))
            colors.append((_LIMIT_COLORS.get(g, {}).get('alert', 'C00000'), False))
        if g in action_map:
            limit_specs.append((f"Grade {g} 조치기준: {action_map[g]:,}", action_map[g]))
            colors.append((_LIMIT_COLORS.get(g, {}).get('action', 'FF3300'), True))


    return _make_chart_sheet(
        wb,
        sheet_name=f"AHU-{ahu_num} {particle_size}",
        chart_title=f"AHU-{ahu_num} {particle_size}µm",
        cat_col_specs=[('청정등급', 'grade'), ('실번호', 'room_num'), ('실명', 'name')],
        data_rows_map=data_rows,
        semesters=semesters,
        limit_specs=limit_specs,
        limit_colors=colors,
    )

# =============================================================================
# B. AIR VELOCITY TEST EXCEL GENERATOR
# =============================================================================

def generate_air_velocity_excel(all_ahu_data, output_path=None):
    """Generate Air Velocity Test Result and Graph Excel file."""
    if output_path is None:
        output_path = os.path.join(OUTPUT_FOLDER, AIR_VELOCITY['excel_filename'])
    wb = Workbook()
    wb.remove(wb.active)
    for ahu_num in sorted(all_ahu_data.keys(), key=lambda x: int(x) if x.isdigit() else x):
        ahu_semesters = all_ahu_data[ahu_num]
        ahu_semesters.sort(key=lambda s: semester_sort_key(s['semester']))
        _create_velocity_data_sheet(wb, ahu_num, ahu_semesters)
        table_ws = _create_velocity_table_sheet(wb, ahu_num, ahu_semesters)
        _create_velocity_chart_sheet(wb, ahu_num, table_ws)
    wb.save(output_path)
    return output_path


def _create_velocity_data_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Data sheet for Air Velocity Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Data")

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = '풍속 측정 기록지'
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN

    # Headers (row 4)
    headers = ['NO.', '청정등급', '실번호', '실명', '측정번호', '측정값 (m/s)', 'Average', '측정일자']
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    current_row = 5
    row_num = 1

    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        machines = sem_data['machines']

        for machine in machines:
            grade = machine['grade']
            room_number = machine['room_number']
            machine_name = machine['machine_name']
            measurements = machine['measurements']
            n_points = len(measurements)

            start_row = current_row

            for i, m in enumerate(measurements):
                r = current_row + i
                ws.cell(row=r, column=1, value=row_num + i)
                apply_cell_style(ws.cell(row=r, column=1))

                if i == 0:
                    ws.cell(row=r, column=2, value=grade)
                    ws.cell(row=r, column=3, value=int(room_number) if room_number.isdigit() else room_number)
                    ws.cell(row=r, column=4, value=machine_name)

                for c in range(2, 5):
                    apply_cell_style(ws.cell(row=r, column=c))

                ws.cell(row=r, column=5, value=m['point'])
                apply_cell_style(ws.cell(row=r, column=5))

                # Velocity value with conditional formatting
                cell_val = ws.cell(row=r, column=6, value=m['value'])
                apply_cell_style(cell_val)
                if m['value'] < AIR_VELOCITY['alert_limits']['low'] or m['value'] > AIR_VELOCITY['alert_limits']['high']:
                    cell_val.fill = ALERT_FILL_RED

                apply_cell_style(ws.cell(row=r, column=7))
                apply_cell_style(ws.cell(row=r, column=8))

            end_row = current_row + n_points - 1

            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)

            # Average
            avg_cell = ws.cell(row=start_row, column=7)
            avg_cell.value = f"=AVERAGE(F{start_row}:F{end_row})"
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

            # 측정일자
            ws.cell(row=start_row, column=8, value=semester_label)
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)

            row_num += n_points
            current_row = end_row + 1

    widths = {'A': 6, 'B': 10, 'C': 10, 'D': 30, 'E': 10, 'F': 14, 'G': 12, 'H': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_velocity_table_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Table sheet for Air Velocity Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Table")

    headers = ['NO.', '청정등급', '실번호', '실명', 'Average', '측정일자']
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    row = 5
    no = 1
    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        for machine in sem_data['machines']:
            measurements = machine['measurements']
            n_points = len(measurements)
            avg = sum(m['value'] for m in measurements) / n_points if n_points > 0 else 0

            ws.cell(row=row, column=1, value=no)
            ws.cell(row=row, column=2, value=machine['grade'])
            ws.cell(row=row, column=3, value=int(machine['room_number']) if machine['room_number'].isdigit() else machine['room_number'])
            ws.cell(row=row, column=4, value=machine['machine_name'])
            ws.cell(row=row, column=5, value=round(avg, 4))
            ws.cell(row=row, column=6, value=semester_label)

            for c in range(1, 7):
                apply_cell_style(ws.cell(row=row, column=c))

            no += 1
            row += 1

    widths = {'A': 6, 'B': 10, 'C': 10, 'D': 30, 'E': 12, 'F': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_velocity_chart_sheet(wb, ahu_num, table_ws):
    """Create AHU-X Pivot chart sheet for Air Velocity Test with limit lines."""
    data_rows = []
    for r in range(5, table_ws.max_row + 1):
        grade    = str(table_ws.cell(row=r, column=2).value or '')
        room_num = table_ws.cell(row=r, column=3).value
        name     = table_ws.cell(row=r, column=4).value
        value    = table_ws.cell(row=r, column=5).value
        semester = table_ws.cell(row=r, column=6).value
        if name is None:
            continue
        data_rows.append({'grade': grade, 'room_num': room_num, 'name': name,
                           'value': value, 'semester': semester})

    semesters = sorted({d['semester'] for d in data_rows if d['semester']}, key=semester_sort_key)
    lo = AIR_VELOCITY['alert_limits']['low']
    hi = AIR_VELOCITY['alert_limits']['high']
    limit_specs = [(f"Lower Limit: {lo} m/s", lo), (f"Upper Limit: {hi} m/s", hi)]
    colors = [(_LIMIT_COLORS['lower'], False), (_LIMIT_COLORS['upper'], False)]


    return _make_chart_sheet(
        wb,
        sheet_name=f"AHU-{ahu_num} Pivot",
        chart_title=f"AHU-{ahu_num}",
        cat_col_specs=[('청정등급', 'grade'), ('실번호', 'room_num'), ('실명', 'name')],
        data_rows_map=data_rows,
        semesters=semesters,
        limit_specs=limit_specs,
        limit_colors=colors,
        note_text=None,
    )


# =============================================================================
# C. AIR CHANGE RATE TEST EXCEL GENERATOR
# =============================================================================

def generate_air_change_rate_excel(all_ahu_data, output_path=None):
    """Generate Air Change Rate Test Result and Graph Excel file."""
    if output_path is None:
        output_path = os.path.join(OUTPUT_FOLDER, AIR_CHANGE_RATE['excel_filename'])
    wb = Workbook()
    wb.remove(wb.active)
    for ahu_num in sorted(all_ahu_data.keys(), key=lambda x: int(x) if x.isdigit() else x):
        ahu_semesters = all_ahu_data[ahu_num]
        ahu_semesters.sort(key=lambda s: semester_sort_key(s['semester']))
        _create_ach_data_sheet(wb, ahu_num, ahu_semesters)
        table_ws = _create_ach_table_sheet(wb, ahu_num, ahu_semesters)
        _create_ach_chart_sheet(wb, ahu_num, table_ws)
    wb.save(output_path)
    return output_path


def _create_ach_data_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Data sheet for Air Change Rate Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Data")

    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = '환기 횟수 측정 결과 기록서'
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN

    # Headers (row 5-6)
    headers_r5 = ['NO', '청정등급', '실번호', '실명', '체적', '측정번호', '측정값', '', '측정일자']
    headers_r6 = ['', '', '', '', '', '', '풍량 (m³/hr)', '환기횟수 (회/hr)', '']

    for col_idx, val in enumerate(headers_r5, 1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    for col_idx, val in enumerate(headers_r6, 1):
        cell = ws.cell(row=6, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    # Merge headers
    ws.merge_cells('G5:H5')
    for c in [1, 2, 3, 4, 5, 6, 9]:
        ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)

    current_row = 7
    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        rooms = sem_data['rooms']

        for room in rooms:
            grade = room['grade']
            room_number = room['room_number']
            room_name = room['room_name']
            volume = room['volume']
            air_flow_measurements = room.get('air_flow_measurements', [])
            total_air_flow = room.get('total_air_flow', 0)
            ach = room['ach']
            n_points = len(air_flow_measurements)

            start_row = current_row

            # Write air flow measurement rows
            for i, af in enumerate(air_flow_measurements):
                r = current_row + i
                if i == 0:
                    ws.cell(row=r, column=1, value=room['no'])
                    ws.cell(row=r, column=2, value=grade)
                    ws.cell(row=r, column=3, value=int(room_number) if room_number.isdigit() else room_number)
                    ws.cell(row=r, column=4, value=room_name)
                    ws.cell(row=r, column=5, value=volume)

                ws.cell(row=r, column=6, value=af['point'])
                ws.cell(row=r, column=7, value=af['air_flow'])

                for c in range(1, 10):
                    apply_cell_style(ws.cell(row=r, column=c))

            # Add 합계 row if multiple points
            if n_points > 1:
                r = current_row + n_points
                ws.cell(row=r, column=6, value='합계')
                ws.cell(row=r, column=7, value=total_air_flow)
                for c in range(1, 10):
                    apply_cell_style(ws.cell(row=r, column=c))
                end_row = r
            else:
                end_row = current_row + n_points - 1

            # ACH value (환기횟수)
            ach_cell = ws.cell(row=start_row, column=8, value=ach)
            apply_cell_style(ach_cell)
            # Conditional formatting
            if grade in AIR_CHANGE_RATE['alert_limits']:
                if ach < AIR_CHANGE_RATE['alert_limits'][grade]:
                    ach_cell.fill = ALERT_FILL_RED

            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
                ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)
                ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)

            # 측정일자
            ws.cell(row=start_row, column=9, value=semester_label)
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9)

            current_row = end_row + 1

    widths = {'A': 6, 'B': 10, 'C': 10, 'D': 16, 'E': 8, 'F': 10, 'G': 16, 'H': 18, 'I': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_ach_table_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Table sheet for Air Change Rate Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Table")

    headers = ['NO', '청정등급', '실번호', '실명', '환기횟수 (회/hr)', '측정일자']
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    row = 5
    no = 1
    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        for room in sem_data['rooms']:
            ws.cell(row=row, column=1, value=no)
            ws.cell(row=row, column=2, value=room['grade'])
            ws.cell(row=row, column=3, value=int(room['room_number']) if room['room_number'].isdigit() else room['room_number'])
            ws.cell(row=row, column=4, value=room['room_name'])
            ws.cell(row=row, column=5, value=room['ach'])
            ws.cell(row=row, column=6, value=semester_label)

            for c in range(1, 7):
                apply_cell_style(ws.cell(row=row, column=c))

            no += 1
            row += 1

    widths = {'A': 6, 'B': 10, 'C': 10, 'D': 16, 'E': 18, 'F': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_ach_chart_sheet(wb, ahu_num, table_ws):
    """Create AHU-X Pivot chart sheet for Air Change Rate Test with limit lines."""
    data_rows = []
    for r in range(5, table_ws.max_row + 1):
        grade    = str(table_ws.cell(row=r, column=2).value or '')
        room_num = table_ws.cell(row=r, column=3).value
        name     = table_ws.cell(row=r, column=4).value
        value    = table_ws.cell(row=r, column=5).value
        semester = table_ws.cell(row=r, column=6).value
        if name is None:
            continue
        data_rows.append({'grade': grade, 'room_num': room_num, 'name': name,
                           'value': value, 'semester': semester})

    semesters = sorted({d['semester'] for d in data_rows if d['semester']}, key=semester_sort_key)
    grades_present = sorted({d['grade'] for d in data_rows if d['grade']})

    alert_map  = AIR_CHANGE_RATE['alert_limits']
    action_map = AIR_CHANGE_RATE.get('action_limits', {})
    limit_specs, colors = [], []
    for g in grades_present:
        if g in alert_map:
            limit_specs.append((f"Grade {g} 경고기준: {alert_map[g]}", alert_map[g]))
            colors.append((_LIMIT_COLORS.get(g, {}).get('alert', 'C00000'), False))
        if g in action_map:
            limit_specs.append((f"Grade {g} 조치기준: {action_map[g]}", action_map[g]))
            colors.append((_LIMIT_COLORS.get(g, {}).get('action', 'FF3300'), True))


    return _make_chart_sheet(
        wb,
        sheet_name=f"AHU-{ahu_num} Pivot",
        chart_title=f"AHU-{ahu_num}",
        cat_col_specs=[('청정등급', 'grade'), ('실번호', 'room_num'), ('실명', 'name')],
        data_rows_map=data_rows,
        semesters=semesters,
        limit_specs=limit_specs,
        limit_colors=colors,
    )
# =============================================================================
# D. HEPA FILTER TEST EXCEL GENERATOR
# =============================================================================

def generate_hepa_filter_excel(all_ahu_data, output_path=None):
    """Generate HEPA Filter Test Result and Graph Excel file."""
    if output_path is None:
        output_path = os.path.join(OUTPUT_FOLDER, HEPA_FILTER['excel_filename'])
    wb = Workbook()
    wb.remove(wb.active)
    for ahu_num in sorted(all_ahu_data.keys(), key=lambda x: int(x) if x.isdigit() else x):
        ahu_semesters = all_ahu_data[ahu_num]
        ahu_semesters.sort(key=lambda s: semester_sort_key(s['semester']))
        _create_hepa_data_sheet(wb, ahu_num, ahu_semesters)
        table_ws = _create_hepa_table_sheet(wb, ahu_num, ahu_semesters)
        _create_hepa_chart_sheet(wb, ahu_num, table_ws)
    wb.save(output_path)
    return output_path


def _create_hepa_data_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Data sheet for HEPA Filter Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Data")

    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'HEPA FILTER 성능 검사 집계표'
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN

    # Headers (row 3)
    headers = ['NO', '실번호', '실명', '측정번호', '측정값', 'Average', '측정일자']
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    current_row = 4
    row_num = 1

    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        items = sem_data['items']

        for item in items:
            room_number = item['room_number']
            item_name = item['item_name']
            measurements = item['measurements']
            n_points = len(measurements)

            start_row = current_row

            for i, m in enumerate(measurements):
                r = current_row + i
                ws.cell(row=r, column=1, value=row_num + i)
                apply_cell_style(ws.cell(row=r, column=1))

                if i == 0:
                    ws.cell(row=r, column=2, value=int(room_number) if room_number.isdigit() else room_number)
                    ws.cell(row=r, column=3, value=item_name)

                for c in range(2, 4):
                    apply_cell_style(ws.cell(row=r, column=c))

                ws.cell(row=r, column=4, value=m['point'])
                apply_cell_style(ws.cell(row=r, column=4))

                # Value as percentage
                val_pct = m['value'] / 100  # Convert from display % to decimal
                cell_val = ws.cell(row=r, column=5, value=val_pct)
                cell_val.number_format = '0.000%'
                apply_cell_style(cell_val)
                if m['value'] > 0.01:  # > 0.01%
                    cell_val.fill = ALERT_FILL_RED

                apply_cell_style(ws.cell(row=r, column=6))
                apply_cell_style(ws.cell(row=r, column=7))

            end_row = current_row + n_points - 1

            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

            # Average
            avg_cell = ws.cell(row=start_row, column=6)
            avg_cell.value = f"=AVERAGE(E{start_row}:E{end_row})"
            avg_cell.number_format = '0.0000%'
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)

            # 측정일자
            ws.cell(row=start_row, column=7, value=semester_label)
            if n_points > 1:
                ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

            row_num += n_points
            current_row = end_row + 1

    widths = {'A': 6, 'B': 10, 'C': 22, 'D': 10, 'E': 12, 'F': 12, 'G': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_hepa_table_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X Table sheet for HEPA Filter Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num} Table")

    headers = ['NO', '실번호', '실명', 'Average', '측정일자']
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    row = 4
    no = 1
    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        for item in sem_data['items']:
            measurements = item['measurements']
            n_points = len(measurements)
            avg = sum(m['value'] for m in measurements) / n_points if n_points > 0 else 0
            avg_decimal = avg / 100  # Convert to decimal for percentage format

            ws.cell(row=row, column=1, value=no)
            ws.cell(row=row, column=2, value=int(item['room_number']) if item['room_number'].isdigit() else item['room_number'])
            ws.cell(row=row, column=3, value=item['item_name'])
            cell_avg = ws.cell(row=row, column=4, value=avg_decimal)
            cell_avg.number_format = '0.0000%'
            ws.cell(row=row, column=5, value=semester_label)

            for c in range(1, 6):
                apply_cell_style(ws.cell(row=row, column=c))

            no += 1
            row += 1

    widths = {'A': 6, 'B': 10, 'C': 22, 'D': 12, 'E': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


def _create_hepa_chart_sheet(wb, ahu_num, table_ws):
    """Create AHU-X Pivot chart sheet for HEPA Filter Test with limit line."""
    data_rows = []
    for r in range(4, table_ws.max_row + 1):
        room_num  = table_ws.cell(row=r, column=2).value
        name      = table_ws.cell(row=r, column=3).value
        value     = table_ws.cell(row=r, column=4).value
        semester  = table_ws.cell(row=r, column=5).value
        if name is None:
            continue
        data_rows.append({'room_num': room_num, 'name': name,
                           'value': value, 'semester': semester})

    semesters = sorted({d['semester'] for d in data_rows if d['semester']}, key=semester_sort_key)
    lim = HEPA_FILTER['alert_limit']
    limit_specs = [('Limit: 0.01%', lim)]
    colors = [(_LIMIT_COLORS['limit'], True)]


    return _make_chart_sheet(
        wb,
        sheet_name=f"AHU-{ahu_num} Pivot",
        chart_title=f"AHU-{ahu_num}",
        cat_col_specs=[('실번호', 'room_num'), ('실명', 'name')],
        data_rows_map=data_rows,
        semesters=semesters,
        limit_specs=limit_specs,
        limit_colors=colors,
        y_num_fmt='0.0000%',
        note_text=None,
    )

# =============================================================================
# E. AIRFLOW PATTERN TEST EXCEL GENERATOR
# =============================================================================

def generate_airflow_pattern_excel(all_ahu_data, output_path=None):
    """Generate Airflow Pattern Test Result and Graph Excel file."""
    wb = Workbook()
    wb.remove(wb.active)

    for ahu_num in sorted(all_ahu_data.keys(), key=lambda x: int(x) if x.isdigit() else x):
        ahu_semesters = all_ahu_data[ahu_num]
        ahu_semesters.sort(key=lambda s: semester_sort_key(s['semester']))
        _create_airflow_sheet(wb, ahu_num, ahu_semesters)

    if output_path is None:
        output_path = os.path.join(OUTPUT_FOLDER, AIRFLOW_PATTERN['excel_filename'])
    wb.save(output_path)
    return output_path


def _create_airflow_sheet(wb, ahu_num, ahu_semesters):
    """Create AHU-X sheet for Airflow Pattern Test."""
    ws = wb.create_sheet(title=f"AHU-{ahu_num}")

    # Headers (row 4)
    headers = ['NO', '측정대상', '측정기준', '동영상 첨부', '판정결과', '측정일자']
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL)

    row = 5
    no = 1
    for sem_data in ahu_semesters:
        semester_label = sem_data['semester']
        items = sem_data['items']

        for item in items:
            ws.cell(row=row, column=1, value=no)
            ws.cell(row=row, column=2, value=item['name'])
            ws.cell(row=row, column=3, value=item['criteria'])
            ws.cell(row=row, column=4, value=item['video_attached'])

            # Judgment with conditional formatting
            judgment_cell = ws.cell(row=row, column=5, value=item['judgment'])
            if item['judgment'] != AIRFLOW_PATTERN['pass_value']:
                judgment_cell.fill = ALERT_FILL_RED

            ws.cell(row=row, column=6, value=semester_label)

            for c in range(1, 7):
                apply_cell_style(ws.cell(row=row, column=c))

            # Set row height for wrapped text in criteria column
            ws.row_dimensions[row].height = 45

            no += 1
            row += 1

    widths = {'A': 6, 'B': 22, 'C': 45, 'D': 12, 'E': 12, 'F': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


# =============================================================================
# MASTER GENERATOR MAP
# =============================================================================

GENERATORS = {
    'airborne_particle': generate_airborne_particle_excel,
    'air_velocity': generate_air_velocity_excel,
    'air_change_rate': generate_air_change_rate_excel,
    'hepa_filter': generate_hepa_filter_excel,
    'airflow_pattern': generate_airflow_pattern_excel,
}
