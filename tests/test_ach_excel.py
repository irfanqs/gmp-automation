import os
import tempfile
import unittest

from openpyxl import load_workbook

from excel_generator import generate_air_change_rate_excel


class AirChangeRateExcelTest(unittest.TestCase):
    def test_table_and_legend_match_required_grade_limits(self):
        rooms = [
            {
                'no': index,
                'grade': grade,
                'room_number': room_number,
                'room_name': room_name,
                'volume': 50.0,
                'air_flow_measurements': [{'point': 1, 'air_flow': 500.0}],
                'total_air_flow': 500.0,
                'ach': ach,
            }
            for index, (grade, room_number, room_name, ach) in enumerate(
                (
                    ('B', '2142', '무균시험실', 63),
                    ('C', '2144', '미생물실험실', 23),
                    ('D', '2165', '균주접종실', 15),
                ),
                start=1,
            )
        ]
        data = {
            '37': [{
                'semester': '2025 (하)',
                'date': '2025.08.02',
                'rooms': rooms,
            }]
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, 'ach.xlsx')
            generate_air_change_rate_excel(data, path)
            workbook = load_workbook(path, data_only=False)

        table = workbook['AHU-37 Table']
        self.assertEqual(
            [table.cell(4, col).value for col in range(1, 12)],
            [
                '청정등급', '실번호', '실명', '환기횟수 (회/hr)', '측정일자',
                'B Grade 경고기준', 'B Grade 조치기준',
                'C Grade 경고기준', 'C Grade 조치기준',
                'D Grade 경고기준', 'D Grade 조치기준',
            ],
        )
        for row in range(5, 8):
            self.assertEqual(
                [table.cell(row, col).value for col in range(6, 12)],
                [52, 50, 22, 20, 12, 10],
            )

        pivot = workbook['AHU-37 Pivot']
        self.assertEqual(pivot['D1'].value, '2025 (하) -')
        self.assertEqual(
            [pivot.cell(1, col).value for col in range(5, 9)],
            [
                '2025 (하) - B Grade 경고기준\n= 52',
                '2025 (하) - B Grade 조치기준\n= 50',
                '2025 (하) - C Grade 경고기준\n= 22',
                '2025 (하) - C Grade 조치기준\n= 20',
            ],
        )
        self.assertEqual(pivot['E2'].value, "='AHU-37 Table'!$F$5")
        self.assertEqual(pivot['H2'].value, "='AHU-37 Table'!$I$5")


if __name__ == '__main__':
    unittest.main()
